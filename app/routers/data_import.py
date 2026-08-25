import csv
import io
import logging
import re
import unicodedata
import zipfile
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func, inspect
from sqlalchemy.orm import Session

from ..api.deps import get_current_active_admin
from ..core.rate_limit import limiter
from ..database import get_db
from ..models import (
    Cliente,
    EmpresaEstadisticas,
    InventarioHistorico,
    InventarioSnapshot,
    Producto,
    ProductoMetricas,
    Registro_PO,
    Usuario,
    VentaHistorica,
)


router = APIRouter(prefix="/data-import", tags=["data-import"])
logger = logging.getLogger(__name__)

MAX_IMPORT_FILE_SIZE = 50 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_SIZE = 512 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200
MAX_XLSX_ARCHIVE_ENTRIES = 2_000
MAX_IMPORT_ROWS = 100_000
MAX_IMPORT_COLUMNS = 64
MAX_REPORTED_ERRORS = 100
MAX_REPORTED_WARNINGS = 100
SUPPORTED_DATASETS = {"products", "inventory", "sales"}
POWER_BI_TRAILER_PREFIXES = ("total", "subtotal", "filtros_aplicados", "filters_applied")
PERCENTAGE_LOSS_FLOOR = -200.0

DATASET_CONFIG = {
    "products": {
        "required": {"sku", "nombre", "costo_unitario", "precio_venta"},
        "headers": [
            "sku",
            "nombre",
            "costo_unitario",
            "precio_venta",
            "lead_time_dias",
            "part_number",
            "ean",
            "peso",
            "familia",
            "marca",
            "familia_marca",
            "product_manager",
            "seccion",
        ],
        "sample": [
            "SKU-001",
            "Producto de ejemplo",
            "12.50",
            "24.90",
            "7",
            "PN-001",
            "8430000000001",
            "0.45",
            "Familia A",
            "Marca A",
            "Familia A/Marca A",
            "Responsable",
            "Seccion A",
        ],
        "aliases": {
            "codigo": "sku",
            "codigo_articulo": "sku",
            "cod_art": "sku",
            "articulo": "sku",
            "nombre_art": "nombre",
            "descripcion": "nombre",
            "descripcion_articulo": "nombre",
            "producto": "nombre",
            "coste_unitario": "costo_unitario",
            "coste": "costo_unitario",
            "costo": "costo_unitario",
            "precio": "precio_venta",
            "pvp": "precio_venta",
            "precio_unitario": "precio_venta",
            "plazo_entrega": "lead_time_dias",
            "lead_time": "lead_time_dias",
            "pn": "part_number",
            "categoria": "familia",
            "familia_marca": "familia_marca",
            "responsable": "product_manager",
            "seccion_madre": "seccion",
        },
    },
    "inventory": {
        "required": {"fecha_inventario", "nombre", "sku", "marca", "familia_marca", "familia", "seccion", "ean", "product_manager", "inventario_eur", "stock_disponible"},
        "headers": ["Fecha", "Nombre Articulo", "ArticuloPK", "Nombre Marca", "Familia/Marca", "Nombre Familia", "Nombre Seccion", "EAN", "Product Manager", "Inventario", "Unidades Inv"],
        "sample": ["06/08/2026", "Producto de ejemplo", "SKU-001", "Marca A", "Moviles/Marca A", "Moviles", "Telefonia", "8430000000001", "Responsable", "1.250,50 €", "150"],
        "canonical_headers": {"fecha_inventario", "nombre", "sku", "marca", "familia_marca", "familia", "seccion", "ean", "product_manager", "inventario_eur", "stock_disponible"},
        "aliases": {
            "fecha": "fecha_inventario",
            "fecha_inventario": "fecha_inventario",
            "codigo": "sku",
            "codigo_articulo": "sku",
            "cod_art": "sku",
            "articulo": "sku",
            "articulopk": "sku",
            "nombre_articulo": "nombre",
            "nombre_art": "nombre",
            "nombre_marca": "marca",
            "nombre_familia": "familia",
            "nombre_seccion": "seccion",
            "familia_marca": "familia_marca",
            "inventario": "inventario_eur",
            "valor_inventario": "inventario_eur",
            "stock": "stock_disponible",
            "stock_actual": "stock_disponible",
            "unidades": "stock_disponible",
            "unidades_inv": "stock_disponible",
            "inventario_unidades": "stock_disponible",
        },
    },
    "sales": {
        "required": {
            "fecha_venta",
            "ingreso_total",
            "cantidad_vendida",
            "margen_bruto_pct",
            "margen_bruto_eur",
            "margen_destino_pct",
            "margen_destino_eur",
            "nombre",
            "sku",
            "marca",
            "familia_marca",
            "familia",
            "seccion",
            "ean",
            "product_manager",
            "cliente_pk",
            "nombre_cliente",
            "kd",
            "tipo_cliente",
            "comercial_cliente",
            "comercial_factura",
        },
        "headers": [
            "Fecha",
            "Ventas",
            "Unidades Venta",
            "% MG",
            "Margen",
            "% MGD",
            "MGD",
            "Nombre Articulo",
            "ArticuloPK",
            "Nombre Marca",
            "Familia/Marca",
            "Nombre Familia",
            "Nombre Seccion",
            "EAN",
            "Product Manager",
            "ClientePK",
            "Nombre Cliente",
            "Kit Digital",
            "Tipo Cliente",
            "Nombre Comercial",
            "Comercial Factura",
        ],
        "sample": [
            "01/05/2026",
            "749.70",
            "3",
            "24.50%",
            "183.68",
            "18.20%",
            "136.45",
            "Producto de ejemplo",
            "SKU-001",
            "Marca A",
            "Moviles/Marca A",
            "Moviles",
            "Telefonia",
            "8430000000001",
            "Responsable",
            "CLI-001",
            "Cliente de ejemplo",
            "NO",
            "Distribuidor",
            "Comercial asignado",
            "Comercial de la venta",
        ],
        "canonical_headers": {
            "fecha_venta",
            "ingreso_total",
            "cantidad_vendida",
            "margen_bruto_pct",
            "margen_bruto_eur",
            "margen_destino_pct",
            "margen_destino_eur",
            "nombre",
            "sku",
            "marca",
            "familia_marca",
            "familia",
            "seccion",
            "ean",
            "product_manager",
            "precio_unitario",
            "stock_disponible",
            "cliente_pk",
            "nombre_cliente",
            "kd",
            "tipo_cliente",
            "comercial_cliente",
            "comercial_factura",
        },
        "aliases": {
            "fecha": "fecha_venta",
            "fecha_factura": "fecha_venta",
            "fecha_transaccion": "fecha_venta",
            "ventas": "ingreso_total",
            "ventas_eur": "ingreso_total",
            "importe": "ingreso_total",
            "importe_total": "ingreso_total",
            "facturacion": "ingreso_total",
            "unidades_venta": "cantidad_vendida",
            "codigo": "sku",
            "codigo_articulo": "sku",
            "cod_art": "sku",
            "articulo": "sku",
            "articulopk": "sku",
            "cantidad": "cantidad_vendida",
            "unidades": "cantidad_vendida",
            "unidades_vendidas": "cantidad_vendida",
            "precio": "precio_unitario",
            "precio_venta": "precio_unitario",
            "pct_mg": "margen_bruto_pct",
            "porcentaje_mg": "margen_bruto_pct",
            "margen_pct": "margen_bruto_pct",
            "margen": "margen_bruto_eur",
            "pct_mgd": "margen_destino_pct",
            "porcentaje_mgd": "margen_destino_pct",
            "mgd_pct": "margen_destino_pct",
            "mgd": "margen_destino_eur",
            "nombre_articulo": "nombre",
            "nombre_art": "nombre",
            "nombre_marca": "marca",
            "familia_marca": "familia_marca",
            "nombre_familia": "familia",
            "nombre_seccion": "seccion",
            "stock": "stock_disponible",
            "stock_actual": "stock_disponible",
            "cliente": "cliente_pk",
            "clientepk": "cliente_pk",
            "codigo_cliente": "cliente_pk",
            "nombre_cliente": "nombre_cliente",
            "kit_digital": "kd",
            "kd": "kd",
            "tipo_cliente": "tipo_cliente",
            "nombre_comercial": "comercial_cliente",
            "comercial_cliente": "comercial_cliente",
            "comercial_factura": "comercial_factura",
        },
    },
}


def _normalize_header(value: str) -> str:
    is_percentage = "%" in value
    normalized = unicodedata.normalize("NFKD", value.strip().lower())
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return f"pct_{normalized}" if is_percentage and not normalized.startswith("pct_") else normalized


def _decode_csv(content: bytes) -> tuple[str, str]:
    try:
        return content.decode("utf-8-sig"), "utf-8"
    except UnicodeDecodeError:
        try:
            return content.decode("cp1252"), "windows-1252"
        except UnicodeDecodeError as exc:
            raise HTTPException(
                status_code=400,
                detail="No se pudo leer el archivo. Guardalo como CSV UTF-8.",
            ) from exc


def _parse_number(value: Any, field: str, required: bool = True) -> float | None:
    raw = str(value or "").strip()
    if not raw:
        if required:
            raise ValueError(f"{field} es obligatorio")
        return None
    cleaned = raw.replace("\u00a0", "").replace("EUR", "").replace("€", "").replace(" ", "")
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    elif re.fullmatch(r"[+-]?\d{1,3}(?:\.\d{3})+", cleaned):
        cleaned = cleaned.replace(".", "")
    try:
        return float(cleaned)
    except ValueError as exc:
        raise ValueError(f"{field} debe ser numerico") from exc


def _parse_integer(value: Any, field: str, required: bool = True) -> int | None:
    parsed = _parse_number(value, field, required)
    if parsed is None:
        return None
    if not float(parsed).is_integer():
        raise ValueError(f"{field} debe ser un numero entero")
    return int(parsed)


def _parse_percentage(value: Any, field: str, required: bool = False) -> float | None:
    raw = str(value or "").strip()
    if not raw:
        if required:
            raise ValueError(f"{field} es obligatorio")
        return None
    has_percent_sign = "%" in raw
    parsed = _parse_number(raw.replace("%", ""), field, required=True)
    if parsed is None:
        return None
    if not has_percent_sign and abs(parsed) <= 1:
        parsed *= 100
    return max(parsed, PERCENTAGE_LOSS_FLOOR)


def _margin_percentage_with_loss_floor(margen_eur: float, ventas_eur: float) -> float | None:
    """Evita porcentajes infinitos y conserva una señal visible para pérdidas sin venta."""
    if abs(ventas_eur) < 0.01:
        return PERCENTAGE_LOSS_FLOOR if margen_eur < 0 else None
    percentage = margen_eur / ventas_eur * 100
    return max(percentage, PERCENTAGE_LOSS_FLOOR)


def _parse_date(value: Any, field: str = "fecha") -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"{field} debe usar YYYY-MM-DD o DD/MM/YYYY")


def _required_text(value: Any, field: str, max_length: int = 255) -> str:
    parsed = str(value or "").strip()
    if not parsed:
        raise ValueError(f"{field} es obligatorio")
    if len(parsed) > max_length:
        raise ValueError(f"{field} supera {max_length} caracteres")
    return parsed


def _optional_text(value: Any, field: str, max_length: int = 255) -> str | None:
    parsed = str(value or "").strip()
    if not parsed:
        return None
    if len(parsed) > max_length:
        raise ValueError(f"{field} supera {max_length} caracteres")
    return parsed


def _canonicalize_headers(headers: list[Any], dataset: str) -> list[str]:
    config = DATASET_CONFIG[dataset]
    aliases = config["aliases"]
    normalized_headers = []
    for header in headers:
        base = _normalize_header(str(header or ""))
        normalized_headers.append(aliases.get(base, base))
    if any(not header for header in normalized_headers):
        raise HTTPException(status_code=400, detail="Hay una columna sin nombre en la cabecera.")
    if len(normalized_headers) > MAX_IMPORT_COLUMNS:
        raise HTTPException(status_code=400, detail=f"El archivo supera el limite de {MAX_IMPORT_COLUMNS} columnas.")
    if len(set(normalized_headers)) != len(normalized_headers):
        raise HTTPException(status_code=400, detail="Hay columnas duplicadas despues de normalizar las cabeceras.")
    missing = sorted(config["required"] - set(normalized_headers))
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas obligatorias: {', '.join(missing)}.",
        )
    return normalized_headers


def _import_metadata(dataset: str, normalized_headers: list[str], **metadata: Any) -> dict[str, Any]:
    config = DATASET_CONFIG[dataset]
    known_headers = set(config.get("canonical_headers", config["headers"]))
    return {
        **metadata,
        "columns": normalized_headers,
        "unknown_columns": sorted(set(normalized_headers) - known_headers),
    }


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_power_bi_trailer(row: dict[str, str]) -> bool:
    """Detecta totales, subtotales y pie de filtros exportados por Power BI."""
    first_value = next((str(value).strip() for value in row.values() if str(value).strip()), "")
    if not first_value:
        return False
    normalized = _normalize_header(first_value)
    return normalized.startswith(POWER_BI_TRAILER_PREFIXES)


def _validate_xlsx_archive(content: bytes) -> None:
    """Permite libros grandes y bloquea patrones propios de una bomba ZIP."""
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        members = archive.infolist()
        if len(members) > MAX_XLSX_ARCHIVE_ENTRIES:
            raise HTTPException(
                status_code=400,
                detail="El XLSX contiene demasiados componentes internos para procesarlo de forma segura.",
            )
        total_uncompressed = sum(item.file_size for item in members)
        if total_uncompressed > MAX_XLSX_UNCOMPRESSED_SIZE:
            raise HTTPException(
                status_code=400,
                detail="El contenido interno del XLSX supera el límite de 512 MB.",
            )
        total_compressed = sum(item.compress_size for item in members)
        compression_ratio = total_uncompressed / max(total_compressed, 1)
        if compression_ratio > MAX_XLSX_COMPRESSION_RATIO:
            raise HTTPException(
                status_code=400,
                detail="El XLSX presenta una compresión anómala y no puede procesarse de forma segura.",
            )


def _read_xlsx(content: bytes, dataset: str) -> tuple[list[dict[str, str]], dict[str, Any]]:
    try:
        _validate_xlsx_archive(content)
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)
    except (zipfile.BadZipFile, InvalidFileException, OSError) as exc:
        raise HTTPException(status_code=400, detail="El archivo XLSX no es valido.") from exc

    try:
        sheet = workbook.active
        rows_iterator = sheet.iter_rows(values_only=True)
        headers = next(rows_iterator, None)
        if not headers:
            raise HTTPException(status_code=400, detail="El XLSX no contiene cabeceras.")
        normalized_headers = _canonicalize_headers(list(headers), dataset)
        rows: list[dict[str, str]] = []
        ignored_powerbi_rows = 0
        reached_filters_footer = False
        for source_row in rows_iterator:
            if len(rows) >= MAX_IMPORT_ROWS:
                raise HTTPException(status_code=400, detail="El XLSX supera el limite de 100.000 filas.")
            normalized_row = {
                header: _cell_to_text(source_row[index] if index < len(source_row) else None)
                for index, header in enumerate(normalized_headers)
            }
            if reached_filters_footer:
                if any(normalized_row.values()):
                    ignored_powerbi_rows += 1
                continue
            if _is_power_bi_trailer(normalized_row):
                ignored_powerbi_rows += 1
                first_value = next((value for value in normalized_row.values() if value), "")
                reached_filters_footer = _normalize_header(first_value).startswith(("filtros_aplicados", "filters_applied"))
                continue
            if any(normalized_row.values()):
                rows.append(normalized_row)
        if not rows:
            raise HTTPException(status_code=400, detail="El XLSX no contiene filas de datos.")
        return rows, _import_metadata(
            dataset,
            normalized_headers,
            encoding="xlsx",
            delimiter=f"hoja: {sheet.title}",
            ignored_powerbi_rows=ignored_powerbi_rows,
        )
    finally:
        workbook.close()


def _read_csv(content: bytes, dataset: str) -> tuple[list[dict[str, str]], dict[str, Any]]:
    decoded, encoding = _decode_csv(content)
    try:
        dialect = csv.Sniffer().sniff(decoded[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(decoded), dialect=dialect)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="El CSV no contiene cabeceras.")
    normalized_headers = _canonicalize_headers(list(reader.fieldnames), dataset)
    rows: list[dict[str, str]] = []
    ignored_powerbi_rows = 0
    reached_filters_footer = False
    for source_row in reader:
        if len(rows) >= MAX_IMPORT_ROWS:
            raise HTTPException(status_code=400, detail="El CSV supera el limite de 100.000 filas.")
        normalized_row = {
            normalized_headers[index]: str(source_row.get(original_header) or "").strip()
            for index, original_header in enumerate(reader.fieldnames)
        }
        if reached_filters_footer:
            if any(normalized_row.values()):
                ignored_powerbi_rows += 1
            continue
        if _is_power_bi_trailer(normalized_row):
            ignored_powerbi_rows += 1
            first_value = next((value for value in normalized_row.values() if value), "")
            reached_filters_footer = _normalize_header(first_value).startswith(("filtros_aplicados", "filters_applied"))
            continue
        if any(normalized_row.values()):
            rows.append(normalized_row)
    if not rows:
        raise HTTPException(status_code=400, detail="El CSV no contiene filas de datos.")
    return rows, _import_metadata(
        dataset,
        normalized_headers,
        encoding=encoding,
        delimiter=dialect.delimiter,
        ignored_powerbi_rows=ignored_powerbi_rows,
    )


async def _read_tabular_file(file: UploadFile, dataset: str) -> tuple[list[dict[str, str]], dict[str, Any]]:
    if dataset not in SUPPORTED_DATASETS:
        raise HTTPException(status_code=400, detail="Tipo de datos no soportado.")
    extension = (file.filename or "").lower().rsplit(".", 1)[-1]
    if extension not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="El archivo debe tener extension .csv o .xlsx.")

    content = await file.read(MAX_IMPORT_FILE_SIZE + 1)
    if len(content) > MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=400, detail="El archivo supera el límite de 50 MB.")
    return _read_xlsx(content, dataset) if extension == "xlsx" else _read_csv(content, dataset)


def _validate_rows(
    dataset: str,
    rows: list[dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    valid_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    seen_keys: set[tuple[str, date] | str] = set()

    for index, row in enumerate(rows, start=2):
        try:
            sku = _required_text(row.get("sku"), "sku", 120)
            if dataset == "products":
                normalized_sku = sku.casefold()
                if normalized_sku in seen_keys:
                    raise ValueError("sku esta duplicado dentro del archivo")
                seen_keys.add(normalized_sku)

            if dataset == "products":
                costo = _parse_number(row.get("costo_unitario"), "costo_unitario")
                precio = _parse_number(row.get("precio_venta"), "precio_venta")
                lead_time = _parse_integer(row.get("lead_time_dias"), "lead_time_dias", required=False)
                peso = _parse_number(row.get("peso"), "peso", required=False)
                if costo is None or costo < 0:
                    raise ValueError("costo_unitario no puede ser negativo")
                if precio is None or precio < 0:
                    raise ValueError("precio_venta no puede ser negativo")
                if lead_time is not None and lead_time < 0:
                    raise ValueError("lead_time_dias no puede ser negativo")
                if peso is not None and peso < 0:
                    raise ValueError("peso no puede ser negativo")
                valid_rows.append({
                    "sku": sku,
                    "nombre": _required_text(row.get("nombre"), "nombre", 255),
                    "costo_unitario": costo,
                    "precio_venta": precio,
                    "lead_time_dias": lead_time if lead_time is not None else 7,
                    "part_number": str(row.get("part_number") or "").strip() or None,
                    "ean": str(row.get("ean") or "").strip() or None,
                    "peso": peso,
                    "familia": str(row.get("familia") or "").strip() or None,
                    "marca": str(row.get("marca") or "").strip() or None,
                    "familia_marca": str(row.get("familia_marca") or "").strip() or None,
                    "product_manager": str(row.get("product_manager") or "").strip() or None,
                    "seccion": str(row.get("seccion") or "").strip() or None,
                })
            elif dataset == "inventory":
                fecha_inventario = _parse_date(row.get("fecha_inventario"), "Fecha")
                inventory_key = (sku.casefold(), fecha_inventario)
                if inventory_key in seen_keys:
                    raise ValueError("la combinación Fecha + ArticuloPK está duplicada dentro del archivo")
                seen_keys.add(inventory_key)
                stock = _parse_integer(row.get("stock_disponible"), "stock_disponible")
                inventario_eur = _parse_number(row.get("inventario_eur"), "Inventario")
                if stock is None or stock < 0:
                    raise ValueError("stock_disponible no puede ser negativo")
                if inventario_eur is None or inventario_eur < 0:
                    raise ValueError("Inventario no puede ser negativo")
                valid_rows.append({
                    "sku": sku,
                    "fecha_inventario": fecha_inventario,
                    "nombre": _required_text(row.get("nombre"), "Nombre Articulo", 255),
                    "marca": _optional_text(row.get("marca"), "Nombre Marca"),
                    "familia_marca": _optional_text(row.get("familia_marca"), "Familia/Marca"),
                    "familia": _optional_text(row.get("familia"), "Nombre Familia"),
                    "seccion": _optional_text(row.get("seccion"), "Nombre Seccion"),
                    "ean": _optional_text(row.get("ean"), "EAN", 80),
                    "product_manager": _optional_text(row.get("product_manager"), "Product Manager"),
                    "inventario_eur": inventario_eur,
                    "stock_disponible": stock,
                })
            else:
                cantidad = _parse_integer(row.get("cantidad_vendida"), "cantidad_vendida")
                precio = _parse_number(row.get("precio_unitario"), "precio_unitario", required=False)
                ingreso = _parse_number(row.get("ingreso_total"), "Ventas")
                margen_bruto_eur = _parse_number(row.get("margen_bruto_eur"), "Margen", required=False)
                margen_bruto_pct = _parse_percentage(row.get("margen_bruto_pct"), "% MG", required=False)
                margen_destino_eur = _parse_number(row.get("margen_destino_eur"), "MGD", required=False)
                margen_destino_pct = _parse_percentage(row.get("margen_destino_pct"), "% MGD", required=False)
                stock = _parse_integer(row.get("stock_disponible"), "stock_disponible", required=False)
                if precio is not None and precio < 0:
                    raise ValueError("precio_unitario no puede ser negativo")
                if stock is not None and stock < 0:
                    raise ValueError("stock_disponible no puede ser negativo")
                if precio is None:
                    precio = float(ingreso) / cantidad if cantidad else 0.0
                if margen_bruto_eur is None and margen_bruto_pct is not None:
                    margen_bruto_eur = float(ingreso) * margen_bruto_pct / 100
                if margen_bruto_pct is None and ingreso and margen_bruto_eur is not None:
                    margen_bruto_pct = margen_bruto_eur / float(ingreso) * 100
                if margen_destino_eur is None and margen_destino_pct is not None:
                    margen_destino_eur = float(ingreso) * margen_destino_pct / 100
                if margen_destino_pct is None and ingreso and margen_destino_eur is not None:
                    margen_destino_pct = margen_destino_eur / float(ingreso) * 100
                valid_rows.append({
                    "sku": sku,
                    "nombre": _required_text(row.get("nombre"), "Nombre Articulo", 255),
                    "fecha_venta": _parse_date(row.get("fecha_venta")),
                    "cantidad_vendida": cantidad or 0,
                    "precio_unitario": precio,
                    "ingreso_total": ingreso,
                    "margen_bruto_eur": margen_bruto_eur or 0.0,
                    "margen_bruto_pct": margen_bruto_pct,
                    "margen_bruto_informado": margen_bruto_eur is not None,
                    "margen_destino_eur": margen_destino_eur or 0.0,
                    "margen_destino_pct": margen_destino_pct,
                    "marca": _optional_text(row.get("marca"), "Nombre Marca"),
                    "familia_marca": _optional_text(row.get("familia_marca"), "Familia/Marca"),
                    "familia": _optional_text(row.get("familia"), "Nombre Familia"),
                    "seccion": _optional_text(row.get("seccion"), "Nombre Seccion"),
                    "ean": _optional_text(row.get("ean"), "EAN", 80),
                    "product_manager": _optional_text(row.get("product_manager"), "Product Manager"),
                    "cliente_pk": _required_text(row.get("cliente_pk"), "ClientePK", 120),
                    "nombre_cliente": _required_text(row.get("nombre_cliente"), "Nombre Cliente", 255),
                    "kd": (_optional_text(row.get("kd"), "KD", 120) or "NO").upper(),
                    "tipo_cliente": _optional_text(row.get("tipo_cliente"), "Tipo Cliente", 120),
                    "comercial_cliente": _optional_text(row.get("comercial_cliente"), "Nombre Comercial", 255),
                    "comercial_factura": _optional_text(row.get("comercial_factura"), "Comercial Factura", 255),
                    "stock_disponible": stock,
                })
        except ValueError as exc:
            if len(errors) < MAX_REPORTED_ERRORS:
                errors.append({"line": index, "message": str(exc)})

    return valid_rows, errors, warnings


def _validation_response(
    dataset: str,
    source_rows: list[dict[str, str]],
    valid_rows: list[dict[str, Any]],
    errors: list[dict[str, Any]],
    warnings: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> dict[str, Any]:
    response = {
        "dataset": dataset,
        "valid": not errors,
        "rows_total": len(source_rows),
        "rows_valid": len(valid_rows),
        "rows_invalid": len(source_rows) - len(valid_rows),
        "errors": errors,
        "warnings": warnings,
        **metadata,
    }
    if dataset in {"sales", "inventory"} and valid_rows:
        date_field = "fecha_venta" if dataset == "sales" else "fecha_inventario"
        dates = [row[date_field] for row in valid_rows]
        response["date_min"] = min(dates).isoformat()
        response["date_max"] = max(dates).isoformat()
    return response


def _company_products(db: Session, empresa_id: int) -> dict[str, Producto]:
    products = db.query(Producto).filter(Producto.empresa_id == empresa_id).all()
    return {product.sku.casefold(): product for product in products}


def _company_clients(db: Session, empresa_id: int) -> dict[str, Cliente]:
    clients = db.query(Cliente).filter(Cliente.empresa_id == empresa_id).all()
    return {client.cliente_pk.casefold(): client for client in clients}


def _clear_company_operational_data(db: Session, empresa_id: int) -> None:
    product_ids = [
        product_id
        for (product_id,) in db.query(Producto.id).filter(Producto.empresa_id == empresa_id).all()
    ]
    if product_ids:
        db.query(VentaHistorica).filter(VentaHistorica.producto_id.in_(product_ids)).delete(synchronize_session=False)
        db.query(Registro_PO).filter(Registro_PO.producto_id.in_(product_ids)).delete(synchronize_session=False)
        db.query(ProductoMetricas).filter(ProductoMetricas.producto_id.in_(product_ids)).delete(synchronize_session=False)
        db.query(InventarioSnapshot).filter(InventarioSnapshot.producto_id.in_(product_ids)).delete(synchronize_session=False)
        db.query(InventarioHistorico).filter(InventarioHistorico.producto_id.in_(product_ids)).delete(synchronize_session=False)
        db.query(Producto).filter(Producto.id.in_(product_ids)).delete(synchronize_session=False)
    db.query(Cliente).filter(Cliente.empresa_id == empresa_id).delete(synchronize_session=False)
    db.query(EmpresaEstadisticas).filter(EmpresaEstadisticas.empresa_id == empresa_id).delete(synchronize_session=False)


def _clear_company_sales_data(db: Session, empresa_id: int) -> None:
    product_ids = [
        product_id
        for (product_id,) in db.query(Producto.id).filter(Producto.empresa_id == empresa_id).all()
    ]
    if product_ids:
        db.query(VentaHistorica).filter(
            VentaHistorica.producto_id.in_(product_ids)
        ).delete(synchronize_session=False)
        db.query(ProductoMetricas).filter(
            ProductoMetricas.producto_id.in_(product_ids)
        ).delete(synchronize_session=False)
    db.query(Cliente).filter(Cliente.empresa_id == empresa_id).delete(synchronize_session=False)
    db.query(EmpresaEstadisticas).filter(
        EmpresaEstadisticas.empresa_id == empresa_id
    ).delete(synchronize_session=False)


def _refresh_metrics(db: Session, empresa_id: int) -> None:
    from ..services import invalidate_metrics_cache, sync_metrics_to_db

    invalidate_metrics_cache(empresa_id)
    sync_metrics_to_db(db, empresa_id)


def _ensure_sales_schema(db: Session) -> None:
    required_columns = {
        "margen_bruto_eur",
        "margen_bruto_pct",
        "margen_destino_eur",
        "margen_destino_pct",
        "cliente_id",
        "kd",
        "comercial_factura",
    }
    required_client_columns = {"empresa_id", "cliente_pk", "nombre", "tipo_cliente", "comercial_cliente"}
    table_names = set(inspect(db.bind).get_table_names())
    sales_columns = {column["name"] for column in inspect(db.bind).get_columns("ventas_historicas")}
    product_columns = {column["name"] for column in inspect(db.bind).get_columns("productos")}
    client_columns = (
        {column["name"] for column in inspect(db.bind).get_columns("clientes")}
        if "clientes" in table_names else set()
    )
    if (
        required_columns - sales_columns
        or required_client_columns - client_columns
        or "familia_marca" not in product_columns
    ):
        raise HTTPException(
            status_code=503,
            detail="La base de datos necesita las migraciones V1.02 y V1.10 antes de cargar fivemin_ventas.",
        )


def _ensure_inventory_schema(db: Session) -> None:
    table_names = set(inspect(db.bind).get_table_names())
    columns = (
        {column["name"] for column in inspect(db.bind).get_columns("inventario_historico")}
        if "inventario_historico" in table_names else set()
    )
    required = {"producto_id", "fecha_inventario", "inventario_eur", "unidades_inventario"}
    if required - columns:
        raise HTTPException(
            status_code=503,
            detail="La base de datos necesita la migración V1.13 antes de cargar el histórico de inventario.",
        )


@router.get("/template/{dataset}")
def download_template(
    dataset: str,
    current_user: Usuario = Depends(get_current_active_admin),
):
    if dataset not in SUPPORTED_DATASETS:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada.")
    config = DATASET_CONFIG[dataset]
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(config["headers"])
    writer.writerow(config["sample"])
    content = output.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="plantilla_{dataset}.csv"'},
    )


@router.post("/validate")
@limiter.limit("10/minute")
async def validate_import(
    request: Request,
    dataset: str = Form(...),
    file: UploadFile = File(...),
    current_user: Usuario = Depends(get_current_active_admin),
):
    source_rows, metadata = await _read_tabular_file(file, dataset)
    valid_rows, errors, warnings = _validate_rows(dataset, source_rows)
    return _validation_response(dataset, source_rows, valid_rows, errors, warnings, metadata)


@router.post("/load")
@limiter.limit("5/minute")
async def load_import(
    request: Request,
    dataset: str = Form(...),
    replace_existing: bool = Form(False),
    sales_mode: str = Form("upsert_keys"),
    file: UploadFile = File(...),
    current_user: Usuario = Depends(get_current_active_admin),
    db: Session = Depends(get_db),
):
    if sales_mode not in {"upsert_keys", "replace_period"}:
        raise HTTPException(status_code=400, detail="Modo de actualizacion de ventas no soportado.")

    source_rows, metadata = await _read_tabular_file(file, dataset)
    valid_rows, errors, warnings = _validate_rows(dataset, source_rows)
    validation = _validation_response(dataset, source_rows, valid_rows, errors, warnings, metadata)
    if errors:
        raise HTTPException(status_code=400, detail=validation)

    empresa_id = current_user.empresa_id
    try:
        if dataset == "sales":
            _ensure_sales_schema(db)
        elif dataset == "inventory":
            _ensure_inventory_schema(db)
        affected = 0
        created = 0
        updated = 0
        products_created = 0
        products_updated = 0
        clients_created = 0
        clients_updated = 0

        if dataset == "products":
            if replace_existing:
                _clear_company_operational_data(db, empresa_id)
                db.flush()
            products = _company_products(db, empresa_id)
            for row in valid_rows:
                product = products.get(row["sku"].casefold())
                if product is None:
                    product = Producto(empresa_id=empresa_id, **row)
                    db.add(product)
                    products[row["sku"].casefold()] = product
                    created += 1
                else:
                    for field, value in row.items():
                        setattr(product, field, value)
                    updated += 1
            affected = len(valid_rows)

        elif dataset == "inventory":
            products = _company_products(db, empresa_id)
            latest_rows: dict[str, dict[str, Any]] = {}
            for row in valid_rows:
                key = row["sku"].casefold()
                if key not in latest_rows or row["fecha_inventario"] >= latest_rows[key]["fecha_inventario"]:
                    latest_rows[key] = row
            for product_key, row in latest_rows.items():
                product = products.get(product_key)
                unit_cost = row["inventario_eur"] / row["stock_disponible"] if row["stock_disponible"] else 0.0
                product_fields = {field: row[field] for field in ("marca", "familia_marca", "familia", "seccion", "ean", "product_manager")}
                if product is None:
                    product = Producto(
                        empresa_id=empresa_id,
                        sku=row["sku"],
                        nombre=row["nombre"],
                        costo_unitario=unit_cost,
                        precio_venta=0.0,
                        lead_time_dias=7,
                        **product_fields,
                    )
                    db.add(product)
                    products[product_key] = product
                    products_created += 1
                else:
                    product.nombre = row["nombre"]
                    product.costo_unitario = unit_cost
                    for field, value in product_fields.items():
                        if value is not None:
                            setattr(product, field, value)
                    products_updated += 1
            db.flush()

            product_ids = list({products[row["sku"].casefold()].id for row in valid_rows})
            incoming_keys = {(products[row["sku"].casefold()].id, row["fecha_inventario"]) for row in valid_rows}
            existing_rows = db.query(InventarioHistorico).filter(
                InventarioHistorico.producto_id.in_(product_ids)
            ).all()
            ids_to_delete = [row.id for row in existing_rows if (row.producto_id, row.fecha_inventario) in incoming_keys]
            if ids_to_delete:
                db.query(InventarioHistorico).filter(InventarioHistorico.id.in_(ids_to_delete)).delete(synchronize_session=False)
                updated = len(ids_to_delete)
            records = [{
                "producto_id": products[row["sku"].casefold()].id,
                "fecha_inventario": row["fecha_inventario"],
                "inventario_eur": row["inventario_eur"],
                "unidades_inventario": row["stock_disponible"],
            } for row in valid_rows]
            db.bulk_insert_mappings(InventarioHistorico, records)
            db.flush()
            latest_inventory: dict[int, InventarioHistorico] = {}
            for item in db.query(InventarioHistorico).filter(
                InventarioHistorico.producto_id.in_(product_ids)
            ).order_by(InventarioHistorico.producto_id, InventarioHistorico.fecha_inventario.desc()).all():
                latest_inventory.setdefault(item.producto_id, item)
            snapshots = {
                snapshot.producto_id: snapshot
                for snapshot in db.query(InventarioSnapshot).filter(InventarioSnapshot.producto_id.in_(product_ids)).all()
            }
            for product_id, item in latest_inventory.items():
                snapshot = snapshots.get(product_id)
                if snapshot is None:
                    db.add(InventarioSnapshot(producto_id=product_id, stock_disponible=item.unidades_inventario))
                else:
                    snapshot.stock_disponible = item.unidades_inventario
            affected = len(records)
            created = max(0, len(records) - updated)

        else:
            if replace_existing:
                _clear_company_sales_data(db, empresa_id)
                db.flush()
            products = _company_products(db, empresa_id)
            clients = _company_clients(db, empresa_id)
            latest_product_rows: dict[str, dict[str, Any]] = {}
            latest_client_rows: dict[str, dict[str, Any]] = {}
            product_totals: dict[str, dict[str, Any]] = {}
            for row in valid_rows:
                product_key = row["sku"].casefold()
                previous = latest_product_rows.get(product_key)
                if previous is None or row["fecha_venta"] >= previous["fecha_venta"]:
                    latest_product_rows[product_key] = row
                totals = product_totals.setdefault(
                    product_key,
                    {"ventas": 0.0, "unidades": 0, "margen": 0.0, "margen_informado": False},
                )
                totals["ventas"] += row["ingreso_total"]
                totals["unidades"] += row["cantidad_vendida"]
                totals["margen"] += row["margen_bruto_eur"]
                totals["margen_informado"] = totals["margen_informado"] or row["margen_bruto_informado"]
                client_key = row["cliente_pk"].casefold()
                previous_client = latest_client_rows.get(client_key)
                if previous_client is None or row["fecha_venta"] >= previous_client["fecha_venta"]:
                    latest_client_rows[client_key] = row

            for product_key, row in latest_product_rows.items():
                product = products.get(product_key)
                totals = product_totals[product_key]
                quantity = totals["unidades"]
                average_price = totals["ventas"] / quantity if quantity else 0.0
                estimated_cost = (
                    (totals["ventas"] - totals["margen"]) / quantity
                    if quantity and totals["margen_informado"]
                    else None
                )
                if estimated_cost is not None and estimated_cost < 0:
                    estimated_cost = None
                product_fields = {
                    "nombre": row["nombre"],
                    "marca": row["marca"],
                    "familia_marca": row["familia_marca"],
                    "familia": row["familia"],
                    "seccion": row["seccion"],
                    "ean": row["ean"],
                    "product_manager": row["product_manager"],
                }
                if product is None:
                    product = Producto(
                        empresa_id=empresa_id,
                        sku=row["sku"],
                        costo_unitario=estimated_cost or 0.0,
                        precio_venta=average_price if average_price >= 0 else 0.0,
                        lead_time_dias=7,
                        **product_fields,
                    )
                    db.add(product)
                    products[product_key] = product
                    products_created += 1
                else:
                    product.nombre = row["nombre"]
                    for field, value in product_fields.items():
                        if field != "nombre" and value is not None:
                            setattr(product, field, value)
                    if average_price >= 0 and quantity:
                        product.precio_venta = average_price
                    if estimated_cost is not None:
                        product.costo_unitario = estimated_cost
                    products_updated += 1
            db.flush()

            for client_key, row in latest_client_rows.items():
                client = clients.get(client_key)
                client_fields = {
                    "nombre": row["nombre_cliente"],
                    "tipo_cliente": row["tipo_cliente"],
                    "comercial_cliente": row["comercial_cliente"],
                }
                if client is None:
                    client = Cliente(
                        empresa_id=empresa_id,
                        cliente_pk=row["cliente_pk"],
                        **client_fields,
                    )
                    db.add(client)
                    clients[client_key] = client
                    clients_created += 1
                else:
                    client.nombre = row["nombre_cliente"]
                    for field, value in client_fields.items():
                        if field != "nombre" and value is not None:
                            setattr(client, field, value)
                    clients_updated += 1
            db.flush()

            inventory_by_product = {
                snapshot.producto_id: snapshot.stock_disponible
                for snapshot in db.query(InventarioSnapshot).join(Producto).filter(
                    Producto.empresa_id == empresa_id
                ).all()
            }
            aggregated: dict[tuple[int, date, int, str, str], dict[str, Any]] = {}
            for row in valid_rows:
                product = products[row["sku"].casefold()]
                client = clients[row["cliente_pk"].casefold()]
                kd = row["kd"] or "NO"
                comercial_factura = row["comercial_factura"] or ""
                key = (product.id, row["fecha_venta"], client.id, kd.casefold(), comercial_factura.casefold())
                item = aggregated.setdefault(key, {
                    "producto_id": product.id,
                    "cliente_id": client.id,
                    "fecha_venta": row["fecha_venta"],
                    "kd": kd,
                    "comercial_factura": row["comercial_factura"],
                    "cantidad_vendida": 0,
                    "ingreso_total": 0.0,
                    "margen_bruto_eur": 0.0,
                    "margen_destino_eur": 0.0,
                    "precio_unitario": row["precio_unitario"],
                    "stock_disponible": row["stock_disponible"],
                })
                item["cantidad_vendida"] += row["cantidad_vendida"]
                item["ingreso_total"] += row["ingreso_total"]
                item["margen_bruto_eur"] += row["margen_bruto_eur"]
                item["margen_destino_eur"] += row["margen_destino_eur"]
                if row["stock_disponible"] is not None:
                    item["stock_disponible"] = row["stock_disponible"]

            if sales_mode == "replace_period":
                date_min = min(row["fecha_venta"] for row in valid_rows)
                date_max = max(row["fecha_venta"] for row in valid_rows)
                product_ids = [product.id for product in products.values()]
                db.query(VentaHistorica).filter(
                    VentaHistorica.producto_id.in_(product_ids),
                    VentaHistorica.fecha_venta.between(date_min, date_max),
                ).delete(synchronize_session=False)
            else:
                product_ids_by_date: dict[date, set[int]] = {}
                for product_id, sale_date, _client_id, _kd, _comercial in aggregated:
                    product_ids_by_date.setdefault(sale_date, set()).add(product_id)
                incoming_keys = set(aggregated)
                for sale_date, product_ids in product_ids_by_date.items():
                    existing_rows = db.query(VentaHistorica).filter(
                        VentaHistorica.producto_id.in_(product_ids),
                        VentaHistorica.fecha_venta == sale_date,
                    ).all()
                    ids_to_delete = [
                        existing.id
                        for existing in existing_rows
                        if existing.cliente_id is None
                        or (
                            existing.producto_id,
                            existing.fecha_venta,
                            existing.cliente_id,
                            (existing.kd or "NO").casefold(),
                            (existing.comercial_factura or "").casefold(),
                        ) in incoming_keys
                    ]
                    if ids_to_delete:
                        db.query(VentaHistorica).filter(
                            VentaHistorica.id.in_(ids_to_delete)
                        ).delete(synchronize_session=False)
                        updated += len(ids_to_delete)

            records = []
            for item in aggregated.values():
                quantity = item["cantidad_vendida"]
                if quantity:
                    item["precio_unitario"] = item["ingreso_total"] / quantity
                item["margen_bruto_pct"] = _margin_percentage_with_loss_floor(
                    item["margen_bruto_eur"], item["ingreso_total"]
                )
                item["margen_destino_pct"] = _margin_percentage_with_loss_floor(
                    item["margen_destino_eur"], item["ingreso_total"]
                )
                item["stock_disponible"] = (
                    item["stock_disponible"]
                    if item["stock_disponible"] is not None
                    else inventory_by_product.get(item["producto_id"], 0)
                )
                records.append(item)
            db.bulk_insert_mappings(VentaHistorica, records)
            affected = len(records)
            created = max(0, len(records) - updated)

        db.flush()
        _refresh_metrics(db, empresa_id)
        return {
            "success": True,
            "dataset": dataset,
            "rows_received": len(source_rows),
            "records_affected": affected,
            "created": created,
            "updated": updated,
            "products_created": products_created,
            "products_updated": products_updated,
            "clients_created": clients_created,
            "clients_updated": clients_updated,
            "replace_existing": replace_existing,
            "sales_mode": sales_mode,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception("Error cargando %s para empresa_id=%s", dataset, empresa_id)
        raise HTTPException(status_code=500, detail="No se pudo completar la carga de datos.") from exc


@router.get("/status")
def import_status(
    current_user: Usuario = Depends(get_current_active_admin),
    db: Session = Depends(get_db),
):
    empresa_id = current_user.empresa_id
    product_count = db.query(func.count(Producto.id)).filter(Producto.empresa_id == empresa_id).scalar() or 0
    inventory_count = db.query(func.count(InventarioSnapshot.producto_id)).join(Producto).filter(
        Producto.empresa_id == empresa_id
    ).scalar() or 0
    inventory_summary = db.query(
        func.count(InventarioHistorico.id),
        func.min(InventarioHistorico.fecha_inventario),
        func.max(InventarioHistorico.fecha_inventario),
    ).join(Producto).filter(Producto.empresa_id == empresa_id).one()
    client_count = db.query(func.count(Cliente.id)).filter(Cliente.empresa_id == empresa_id).scalar() or 0
    sales_summary = db.query(
        func.count(VentaHistorica.id),
        func.min(VentaHistorica.fecha_venta),
        func.max(VentaHistorica.fecha_venta),
    ).join(Producto).filter(Producto.empresa_id == empresa_id).one()
    return {
        "products": int(product_count),
        "inventory_records": int(inventory_count),
        "inventory_history_records": int(inventory_summary[0] or 0),
        "inventory_date_min": inventory_summary[1].isoformat() if inventory_summary[1] else None,
        "inventory_date_max": inventory_summary[2].isoformat() if inventory_summary[2] else None,
        "clients": int(client_count),
        "sales_records": int(sales_summary[0] or 0),
        "sales_date_min": sales_summary[1].isoformat() if sales_summary[1] else None,
        "sales_date_max": sales_summary[2].isoformat() if sales_summary[2] else None,
    }
